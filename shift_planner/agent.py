import os
import datetime
from google.adk.agents import Agent
from google.adk.models import Gemini
from google.genai import types
from dotenv import load_dotenv

from shift_planner.solver import ShiftSolver
from shift_planner.sheets_helper import (
    get_credentials,
    get_or_create_folder,
    get_or_create_spreadsheet,
    write_schedule_to_sheets,
)

# Load environment variables
load_dotenv()

# Setup Gemini model configuration for ADK
model_name = os.environ.get("GEMINI_MODEL_NAME", "gemini-2.5-flash")

# Define tools for the agent

def parse_date(date_str: str) -> datetime.date:
    for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Could not parse date: '{date_str}'. Please use DD-MMM-YYYY (e.g. 06-Jul-2026).")

def generate_shift_schedule(employees: list[str], start_date_str: str, end_date_str: str) -> list[dict]:
    """
    Generates a balanced shift schedule for the specified date range
    satisfying all constraints:
    - 2 off days (weekoffs) per week per employee.
    - Exactly 1 person in Shift1 (07:00-15:00 IST) and 1 person in Shift2 (12:00-20:00 IST) on Sundays.
    - Balanced weekday coverage.

    Args:
        employees: List of employee names.
        start_date_str: Start date of the schedule (e.g., '06-Jul-2026').
        end_date_str: End date of the schedule (e.g., '02-Aug-2026').

    Returns:
        List of daily schedule records containing Date, Day, Employee name, and Assignment.
    """
    start_date = parse_date(start_date_str)
    end_date = parse_date(end_date_str)
    solver = ShiftSolver(employees)
    return solver.generate_schedule(start_date, end_date)

def upload_schedule_to_sheets(employees: list[str], schedule: list[dict]) -> str:
    """
    Attempts to upload the generated schedule to Google Sheets in folder 'KaggleCap'
    under a name based on the schedule dates (e.g., 'Shift_Details_jul_2026').
    If credentials are not present, it saves a local Excel (.xlsx) file and returns details.

    Args:
        employees: List of employee names.
        schedule: List of daily schedule records.

    Returns:
        A string message describing the outcome.
    """
    # 1. Generate local Excel file first
    first_record = schedule[0]
    start_date_obj = datetime.datetime.strptime(first_record["Date"], "%Y-%m-%d").date()
    month_str = start_date_obj.strftime("%b").lower()
    year_str = start_date_obj.strftime("%Y")
    suffix = f"{month_str}_{year_str}"
    
    excel_file = f"Shift_Details_{suffix}.xlsx"
    spreadsheet_name = f"Shift_Details_{suffix}"
    tab_title = start_date_obj.strftime("%B %Y")
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    employees_sorted = sorted(list(set(employees)))
    headers = ["Date", "Day"] + employees_sorted
    
    from collections import defaultdict
    pivoted = defaultdict(dict)
    for r in schedule:
        pivoted[(r["Date"], r["Day"])][r["Employee"]] = r["Assignment"]
        
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = tab_title
        
        # Write headers
        ws.append(headers)
        
        # Write rows
        for (date, day) in sorted(pivoted.keys()):
            row = [date, day]
            for emp in employees_sorted:
                row.append(pivoted[(date, day)].get(emp, ""))
            ws.append(row)
            
        # Styling definitions
        font_family = "Segoe UI"
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(name=font_family, size=11, bold=True)
        regular_font = Font(name=font_family, size=10)
        weekoff_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        weekoff_font = Font(name=font_family, size=10, color="9C0006", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Format Header Row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
        # Format Data Rows
        for row_idx in range(2, len(pivoted) + 2):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = regular_font
                cell.alignment = center_align
                
                # Highlight weekoffs in RED
                if col_idx > 2 and cell.value == "Weekoff":
                    cell.fill = weekoff_fill
                    cell.font = weekoff_font
                    
        # Auto-fit column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        wb.save(excel_file)
    except Exception as e:
        print(f"Error saving local Excel file: {e}")

    # 2. Check credentials
    creds = get_credentials()
    if creds is None:
        return (
            f"Local Excel file saved successfully:\n"
            f"- Excel File: `{excel_file}`\n\n"
            f"**Google Sheets integration was skipped** because no credentials were found.\n"
            f"To enable direct Google Sheets upload, please place either `credentials.json` "
            f"(OAuth Client) or `service_account.json` (Service Account) in the project directory."
        )
        
    try:
        # Build drive service
        from googleapiclient.discovery import build
        drive_service = build("drive", "v3", credentials=creds)
        
        # Get or create folder
        folder_id = get_or_create_folder(drive_service, "KaggleCap")
        
        # Get or create spreadsheet
        spreadsheet_id = get_or_create_spreadsheet(drive_service, folder_id, spreadsheet_name)
        
        # Populate sheet
        write_schedule_to_sheets(creds, spreadsheet_id, employees, schedule)
        
        return (
            f"Successfully uploaded schedule to Google Sheet named '{spreadsheet_name}' "
            f"inside the Google Drive folder 'KaggleCap'.\n"
            f"Local Excel backup also saved as `{excel_file}`."
        )
    except Exception as e:
        return (
            f"Failed to upload to Google Sheets due to an error: {e}.\n"
            f"However, local Excel backup was successfully saved to `{excel_file}`."
        )

# Define the root agent
root_agent = Agent(
    name="shift_planner_agent",
    model=Gemini(
        model=model_name,
        retry_options=types.HttpRetryOptions(attempts=3),
    ),
    instruction=(
        "You are an AI shift planning assistant. Given a list of employee names, a start date, and an end date:\n"
        "1. Run the 'generate_shift_schedule' tool to calculate the shift schedule for the specified date range.\n"
        "2. Run the 'upload_schedule_to_sheets' tool to save it locally and upload it to Google Sheets.\n"
        "3. Present the resulting schedule back to the user in a beautiful format and explain briefly how it satisfies the required constraints:\n"
        "   - Shift1 (07:00 - 15:00 IST) and Shift2 (12:00 - 20:00 IST).\n"
        "   - Each employee gets exactly 2 weekoff days per week.\n"
        "   - On Sundays, exactly 1 employee works Shift1 and 1 works Shift2, and the other employees are on weekoff.\n"
        "   - Weekday coverage is balanced."
    ),
    tools=[generate_shift_schedule, upload_schedule_to_sheets],
)
